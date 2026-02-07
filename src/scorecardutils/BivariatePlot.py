#myfucn
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os
import numpy as np
from openpyxl.drawing.image import Image
from openpyxl import Workbook
import openpyxl
from typing import List, Optional, Union, Dict, Any, Tuple
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# Professional color palette for charts
CHART_COLORS = {
    'train_bar': "#2A6F97",       # Office blue
    'oot_bar': "#E76F51",         # Office orange
    'train_line': "#BB3D4E",      # Emerald green
    'oot_line': '#8E44AD',        # Amethyst purple
    'train_special': '#C0392B',   # Dark red
    'oot_special': '#922B21',     # Darker red
    'train_missing': '#2980B9',   # Steel blue
    'oot_missing': '#1F618D',     # Darker steel blue
    'train_label': '#1E8449',     # Dark green for text
    'oot_label': '#6C3483',       # Dark purple for text
}

def unified_bivariate_analysis(
    binning_process,
    filename: str = 'bivariate_analysis',
    metric: str = 'event_rate',
    variables: Optional[List[str]] = None,
    figsize: tuple = (12, 6),
    dpi: int = 100,
    style: str = 'whitegrid',
    oot_data: Optional[pd.DataFrame] = None,
    target_column: Optional[str] = None,
    compare_data: bool = True,
    show_bar_values: bool = False,
    verbose: bool = False
) -> None:
    """
    Unified function to create bivariate analysis plots for binned data with options for:
    1. Training data only
    2. OOT data only (based on training binning)
    3. Comparison between training and OOT data
    
    Parameters:
    -----------
    binning_process : OptimalBinning process object
        The binning process containing the variables to plot
    filename : str, default='bivariate_analysis'
        Base name for the output Excel file
    metric : str, default='event_rate'
        Metric to plot. Options: 'event_rate', 'woe'
    variables : list of str, optional
        List of specific variable names to process. If None, all variables are processed.
    figsize : tuple, default=(12, 6)
        Figure size for plots in inches (width, height)
    dpi : int, default=100
        Resolution of saved images
    style : str, default='whitegrid'
        Seaborn style for plots
    oot_data : pandas DataFrame, optional
        The OOT dataset to transform and analyze. If None, only training data is plotted.
    target_column : str, optional
        Name of the target/outcome column in the OOT data. Required if oot_data is provided.
    compare_data : bool, default=True
        Whether to plot training data metrics alongside OOT data metrics.
        Only relevant when oot_data is provided.
    show_bar_values : bool, default=False
        Whether to show Count (%) values on top of the bars
    verbose : bool, default=False
        Whether to print detailed information during processing
    """
    # Validate inputs
    if oot_data is not None and target_column is None:
        raise ValueError("target_column must be specified when oot_data is provided")
    
    # Set mode based on inputs
    if oot_data is None:
        mode = "train_only"
    else:
        mode = "oot_comparison" if compare_data else "oot_only"
        
    # Validate metric parameter
    metric = metric.lower()
    if metric not in ['event_rate', 'woe']:
        raise ValueError("metric must be either 'event_rate' or 'woe'")
    
    # Get all available variables from binning process
    try:
        # Try to get variable names directly
        all_variables = binning_process.variable_names
    except AttributeError:
        # Fall back to get_support if variable_names is not available
        try:
            all_variables = binning_process.get_support(names=True)
        except AttributeError:
            raise ValueError("Unable to extract variable names from binning_process")
    
    # Filter variables if specified
    if variables is not None:
        # Check if all specified variables exist in the binning process
        invalid_vars = set(variables) - set(all_variables)
        if invalid_vars:
            raise ValueError(f"Variables not found in binning process: {', '.join(invalid_vars)}")
        selected_vars = variables
    else:
        selected_vars = all_variables
    
    # Track if we've created at least one valid sheet for Excel
    valid_sheets_created = False
    
    # Set up the visualization style
    plt.style.use('default')  # Clean matplotlib style
    
    # Create a temporary directory for images if it doesn't exist
    if not os.path.exists('temp_plots'):
        os.makedirs('temp_plots')
    
    # Dictionary to store plot paths and binning tables
    plot_paths = {}
    oot_binning_tables = {}
    
    # Determine column names based on metric
    metric_column = 'Event rate' if metric == 'event_rate' else 'WoE'
    y_axis_label = 'Event Rate' if metric == 'event_rate' else 'Weight of Evidence (WoE)'
    
    # Create Excel writer
    with pd.ExcelWriter(f'{filename}.xlsx', engine='openpyxl', mode='w') as writer:
        # Process each selected variable
        for var in selected_vars:
            try:
                # Sanitize variable name for Excel sheet name (remove invalid characters)
                sheet_name = str(var)
                for char in ['[', ']', ':', '*', '?', '/', '\\']:
                    sheet_name = sheet_name.replace(char, '_')
                # Ensure sheet name is not longer than 31 characters (Excel limit)
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]
                    
                # Get the binned variable from the trained process
                try:
                    optb = binning_process.get_binned_variable(var)
                except Exception as e:
                    if verbose:
                        print(f"Error getting binned variable {var}: {str(e)}")
                    continue
                
                # Get original binning table from training data
                train_df = optb.binning_table.build()
                if 'JS' in train_df.columns:
                    train_df = train_df.drop(columns='JS')
                
                train_df[['Count (%)', 'Event rate', 'IV']] = (train_df[['Count (%)', 'Event rate', 'IV']] * 100).round(2)
                train_df['WoE'] = pd.to_numeric(train_df['WoE'], errors='coerce')
                train_df['WoE'] = train_df['WoE'].round(4)

                # Keep the original DataFrame for Excel output including Totals row
                train_df_for_excel = train_df.reset_index()
                
                # Add Dataset column to training data
                train_df_for_excel.insert(0, 'Dataset', 'Training')
                
                # Extract the Totals row and prepare for plotting
                train_totals_row = None
                train_df_no_totals = train_df
                if 'Totals' in train_df.index:
                    train_totals_row = train_df.loc['Totals']
                    train_df_no_totals = train_df.drop('Totals')
                
                # Reset index for plotting
                train_df_reset = train_df_no_totals.reset_index()
                
                # The actual bin column is 'Bin', but if it doesn't exist, use the first column as fallback
                bin_col_name = 'Bin' if 'Bin' in train_df_reset.columns else train_df_reset.columns[0]
                
                # Process OOT data if provided
                oot_df = None
                oot_df_with_totals = None
                total_event_rate_oot = None
                total_iv_oot = None
                
                if oot_data is not None:
                    # Transform the OOT data for this variable using the trained binning
                    try:
                        X_var_oot = oot_data[var].values
                        y_oot = oot_data[target_column].values
                    except KeyError as e:
                        if verbose:
                            print(f"Variable {var} or target column {target_column} not found in OOT data: {str(e)}")
                        continue
                    
                    # Use the transform method with "bins" metric to get bin assignments
                    try:
                        oot_bins = optb.transform(X_var_oot, metric="bins")
                    except Exception as e:
                        if verbose:
                            print(f"Error transforming OOT data for variable {var}: {str(e)}")
                        continue
                    
                    # Create a dictionary to store counts for each bin - convert to strings to make hashable
                    bin_n_event = {str(bin_name): 0 for bin_name in train_df_reset[bin_col_name] if str(bin_name).strip() != ''}
                    bin_n_nonevent = {str(bin_name): 0 for bin_name in train_df_reset[bin_col_name] if str(bin_name).strip() != ''}
                    
                    # Count events and non-events for each bin in OOT data
                    for i, bin_label in enumerate(oot_bins):
                        bin_label_str = str(bin_label)
                        if bin_label_str in bin_n_event:
                            if y_oot[i] == 1:
                                bin_n_event[bin_label_str] += 1
                            else:
                                bin_n_nonevent[bin_label_str] += 1
                    
                    # Calculate totals for percentages
                    total_events_oot = sum(bin_n_event.values())
                    total_nonevents_oot = sum(bin_n_nonevent.values())
                    total_records_oot = total_events_oot + total_nonevents_oot
                    
                    # Create a new DataFrame for OOT data metrics
                    oot_stats = []
                    total_iv_oot = 0
                    for bin_name in train_df_reset[bin_col_name]:
                        # Convert bin name to string for consistency
                        if str(bin_name).strip():
                            bin_name_str = str(bin_name)
                            n_event = bin_n_event[bin_name_str]
                            n_nonevent = bin_n_nonevent[bin_name_str]
                            n_records = n_event + n_nonevent
                            
                            # Calculate event rate for this bin
                            event_rate = n_event / n_records if n_records > 0 else 0
                            
                            # Calculate event and non-event percentages
                            event_pct = n_event / total_events_oot if total_events_oot > 0 else 0
                            nonevent_pct = n_nonevent / total_nonevents_oot if total_nonevents_oot > 0 else 0
                            
                            # Calculate WoE similar to the binning_table.build() method
                            if n_event > 0 and n_nonevent > 0 and total_events_oot > 0 and total_nonevents_oot > 0:
                                p_event = n_event / total_events_oot
                                p_nonevent = n_nonevent / total_nonevents_oot
                                woe = np.log(p_nonevent / p_event)
                            else:
                                woe = 0  # Default value
                            
                            # Calculate IV for this bin
                            iv = (nonevent_pct - event_pct) * woe
                            total_iv_oot += iv
                            
                            oot_stats.append({
                                bin_col_name: bin_name,  # Keep as original type for consistency
                                'Count': n_records,
                                'Count (%)': n_records / total_records_oot if total_records_oot > 0 else 0,
                                'Non-event': n_nonevent,
                                'Non-event (%)': nonevent_pct,
                                'Event': n_event,
                                'Event (%)': event_pct,
                                'Event rate': event_rate,
                                'WoE': woe,
                                'IV': iv
                            })
                    
                    # Create DataFrame from collected statistics
                    oot_df = pd.DataFrame(oot_stats)
                    oot_df.reset_index(inplace=True)
                    
                    # Calculate total event rate for OOT
                    total_event_rate_oot = total_events_oot / total_records_oot if total_records_oot > 0 else 0
                    
                    # Add totals row for OOT
                    totals_row_oot = {
                        'index': 'Totals',
                        bin_col_name: '',
                        'Count': total_records_oot,
                        'Count (%)': 1.0,
                        'Non-event': total_nonevents_oot,
                        'Non-event (%)': 1.0,
                        'Event': total_events_oot,
                        'Event (%)': 1.0,
                        'Event rate': total_event_rate_oot,
                        'WoE': '',  # WoE doesn't make sense for totals
                        'IV': total_iv_oot
                    }
                    
                    # Add totals row to OOT data
                    oot_df_with_totals = pd.concat([oot_df, pd.DataFrame([totals_row_oot])], ignore_index=True)
                    
                    oot_df_with_totals[['Count (%)', 'Event rate', 'IV']] = (oot_df_with_totals[['Count (%)', 'Event rate', 'IV']] * 100).round(2)
                    oot_df_with_totals['WoE'] = pd.to_numeric(oot_df_with_totals['WoE'], errors='coerce')
                    oot_df_with_totals['WoE'] = oot_df_with_totals['WoE'].round(4)

                    # Add Dataset column to OOT data
                    oot_df_with_totals.insert(0, 'Dataset', 'OOT')

                    # Extract the Totals row and prepare for plotting
                    oot_totals_row = None
                    oot_df_no_totals = oot_df_with_totals.set_index('index')
                    if 'Totals' in oot_df_no_totals.index:
                        oot_totals_row = oot_df_no_totals.loc['Totals']
                        oot_df_no_totals = oot_df_no_totals.drop('Totals')

                    # Reset index for plotting
                    oot_df_reset = oot_df_no_totals.reset_index()
                    
                    # Store OOT binning table
                    oot_binning_tables[var] = oot_df_with_totals
                

                # Determine which data to write to Excel based on mode
                if mode == "train_only":
                    df_for_excel = train_df_for_excel
                    
                elif mode == "oot_only":
                    df_for_excel = oot_df_with_totals
                    
                else:  # oot_comparison - write both with headers and gap
                    # Ensure both dataframes have the same columns in the same order
                    all_columns = list(set(train_df_for_excel.columns) | set(oot_df_with_totals.columns))
                    
                    # Move 'Dataset' column to the beginning if present
                    if 'Dataset' in all_columns:
                        all_columns.remove('Dataset')
                        all_columns = ['Dataset'] + all_columns
                    
                    # Ensure both dataframes have all columns
                    for col in all_columns:
                        if col not in train_df_for_excel.columns:
                            train_df_for_excel[col] = ''
                        if col not in oot_df_with_totals.columns:
                            oot_df_with_totals[col] = ''
                    
                    # Reorder columns to match
                    train_df_for_excel = train_df_for_excel[all_columns]
                    oot_df_with_totals = oot_df_with_totals[all_columns]
                    
                    # Create a blank row for spacing
                    blank_row = pd.DataFrame([{col: '' for col in all_columns}])
                    
                    
                    # Combine all parts: variable name, training header, training data, blank row, OOT header, OOT data
                    df_for_excel = pd.concat([
                        train_df_for_excel, 
                        blank_row, 
                        oot_df_with_totals
                    ], ignore_index=True)
                
                # Reorder columns to the desired structure
                desired_column_order = [
                    'Dataset', 
                    bin_col_name, 
                    'Count', 
                    'Count (%)', 
                    'Non-event', 
                    #'Non-event (%)', 
                    'Event', 
                    #'Event (%)', 
                    'Event rate', 
                    'WoE', 
                    'IV'
                ]

                # Reorder columns (only for columns that exist in the DataFrame)
                existing_columns = [col for col in desired_column_order if col in df_for_excel.columns]
                df_for_excel = df_for_excel[existing_columns]
                
                # Write to Excel
                df_for_excel.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

                worksheet = writer.sheets[sheet_name]
                num_columns = len(df_for_excel.columns)
                num_rows = len(df_for_excel) + 2

                white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                title_fill = PatternFill(start_color="2A6F97", end_color="2A6F97", fill_type="solid")
                header_fill = PatternFill(start_color="468FAF", end_color="468FAF", fill_type="solid")
                totals_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
                inner_border = Border(
                    left=Side(style='thin', color='BDC3C7'),
                    right=Side(style='thin', color='BDC3C7'),
                    top=Side(style='thin', color='BDC3C7'),
                    bottom=Side(style='thin', color='BDC3C7')
                )
                totals_border = Border(
                    left=Side(style='thin', color='BDC3C7'),
                    right=Side(style='thin', color='BDC3C7'),
                    top=Side(style='medium', color='2A6F97'),
                    bottom=Side(style='medium', color='2A6F97')
                )
                title_font = Font(bold=True, size=13, color="FFFFFF", name='Calibri')
                header_font = Font(bold=True, size=11, color="FFFFFF", name='Calibri')
                data_font = Font(size=11, color="1B2631", name='Calibri')
                totals_font = Font(bold=True, size=11, color="1A5276", name='Calibri')

                # Detect totals and blank separator rows
                totals_excel_rows = set()
                blank_excel_rows = set()
                for i in range(len(df_for_excel)):
                    row_data = df_for_excel.iloc[i]
                    bin_val = str(row_data.get(bin_col_name, '')).strip()
                    if str(row_data.get('Dataset', '')).strip() == '' and bin_val == '' and str(row_data.get('Count', '')).strip() == '':
                        blank_excel_rows.add(i + 3)
                        continue
                    if bin_val == 'Totals':
                        totals_excel_rows.add(i + 3)
                    elif bin_val == '':
                        try:
                            if float(row_data.get('Count (%)', 0)) >= 99.99:
                                totals_excel_rows.add(i + 3)
                        except (ValueError, TypeError):
                            pass

                # White fill for entire visible sheet area (up to column BB)
                for row_num in range(1, max(num_rows + 50, 100)):
                    for col_num in range(1, 55):
                        worksheet.cell(row=row_num, column=col_num).fill = white_fill

                # Title row
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)
                worksheet.cell(row=1, column=1).value = var
                worksheet.row_dimensions[1].height = 30
                for col_num in range(1, num_columns + 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = title_fill
                    cell.font = title_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='medium', color='2A6F97'),
                        right=Side(style='medium', color='2A6F97'),
                        top=Side(style='medium', color='2A6F97'),
                        bottom=Side(style='thin', color='2A6F97')
                    )

                # Column headers
                worksheet.row_dimensions[2].height = 26
                for col_num in range(1, num_columns + 1):
                    cell = worksheet.cell(row=2, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin', color='468FAF'),
                        right=Side(style='thin', color='468FAF'),
                        top=Side(style='thin', color='2A6F97'),
                        bottom=Side(style='medium', color='2A6F97')
                    )

                # Data rows
                for row_num in range(3, num_rows + 1):
                    worksheet.row_dimensions[row_num].height = 20
                    if row_num in blank_excel_rows:
                        worksheet.row_dimensions[row_num].height = 8
                        for col_num in range(1, num_columns + 1):
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell.fill = white_fill
                            cell.border = Border()
                        continue
                    is_totals = row_num in totals_excel_rows
                    for col_num in range(1, num_columns + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.fill = totals_fill if is_totals else white_fill
                        cell.font = totals_font if is_totals else data_font
                        cell.border = totals_border if is_totals else inner_border
                        cell.alignment = Alignment(vertical='center')

                valid_sheets_created = True
                
                # Create plot image path
                plot_image_path = f'temp_plots/variable_{var}_{metric}_{mode}.png'
                
                # Create the appropriate plot based on mode
                fig, ax1 = plt.subplots(figsize=figsize)
                
                if mode == "train_only":
                    # Plot training data only
                    plot_single_dataset(
                        ax1, train_df_reset, bin_col_name, metric_column,
                        y_axis_label, train_totals_row, "Training",
                        line_color=CHART_COLORS['train_line'],
                        bar_color=CHART_COLORS['train_bar'],
                        show_bar_values=show_bar_values
                    )
                    plt.title(f'{var}: {y_axis_label} by Bin (Training Data)', fontsize=12)

                elif mode == "oot_only":
                    # Plot OOT data only
                    plot_single_dataset(
                        ax1, oot_df_reset, bin_col_name, metric_column,
                        y_axis_label, oot_totals_row,
                        "OOT", line_color=CHART_COLORS['oot_line'],
                        bar_color=CHART_COLORS['oot_bar'],
                        show_bar_values=show_bar_values
                    )
                    plt.title(f'{var}: {y_axis_label} by Bin (OOT Data)', fontsize=12)
                    
                else:  # oot_comparison
                    # Plot comparison between training and OOT data
                    plot_comparison(
                        ax1, train_df_reset, oot_df_reset, bin_col_name, metric_column,
                        y_axis_label, train_totals_row, oot_totals_row, metric, show_bar_values=show_bar_values
                    )
                    plt.title(f'{var}: {y_axis_label} Comparison (Training vs OOT)', fontsize=12)
                
                # Adjust layout and save
                plt.tight_layout()
                plt.savefig(plot_image_path, dpi=dpi, bbox_inches='tight')
                plt.close()
                
                # Store path for cleanup
                plot_paths[var] = plot_image_path
                
                # Insert plot image into Excel
                try:
                    img = Image(plot_image_path)
                    
                    # Start plot after the data table with some margin
                    row_position = len(df_for_excel) + 4
                    img.anchor = f'A{row_position}'
                    
                    writer.sheets[sheet_name].add_image(img)
                    
                    # Adjust column widths
                    for idx, col in enumerate(df_for_excel.columns):
                        column_width = max(len(str(col)), df_for_excel[col].astype(str).map(len).max())
                        writer.sheets[sheet_name].column_dimensions[chr(65 + idx)].width = column_width + 4
                except Exception as e:
                    if verbose:
                        print(f"Error adding image to Excel for variable {var}: {str(e)}")
            
            except Exception as e:
                if verbose:
                    print(f"Error processing variable {var}: {str(e)}")
                continue
        
        # Add a dummy sheet if no valid sheets were created to avoid Excel error
        if not valid_sheets_created:
            dummy_df = pd.DataFrame({'Message': ['No valid variables found for analysis']})
            dummy_df.to_excel(writer, sheet_name='Info', index=False)
            if verbose:
                print("No valid sheets were created. Adding a dummy sheet.")
    
    # Clean up temporary image files
    for path in plot_paths.values():
        if os.path.exists(path):
            os.remove(path)
    
    # Remove temp directory if empty
    if os.path.exists('temp_plots') and not os.listdir('temp_plots'):
        os.rmdir('temp_plots')
    
    if verbose:
        print(f"Bivariate analysis completed! Results saved to {filename}.xlsx")
    
    # Return None instead of the oot_binning_tables to avoid printing
    return None


def plot_single_dataset(
    ax1, df, bin_col_name, metric_column, y_axis_label,
    totals_row, dataset_name, line_color='#27AE60', bar_color='#3498DB',
    show_bar_values=False
) -> None:
    """
    Helper function to plot a single dataset (either training or OOT)
    """
    # Ensure bin column is string type
    df[bin_col_name] = df[bin_col_name].astype(str)

    # Identify regular, special, and missing bins
    regular_bins = df[~df[bin_col_name].str.contains('Special|Missing', regex=True, na=False)]
    special_bin = df[df[bin_col_name].str.contains('Special', regex=False, na=False)]
    missing_bin = df[df[bin_col_name].str.contains('Missing', regex=False, na=False)]

    # Create indices for x-axis
    x_indices = np.arange(len(df))

    # Format bin labels
    bin_labels = []
    for _, row in df.iterrows():
        bin_name = row[bin_col_name]
        if 'Special' in bin_name or 'Missing' in bin_name:
            bin_labels.append(bin_name)
        else:
            shortened_name = str(bin_name)
            if len(shortened_name) > 15:
                shortened_name = shortened_name[:12] + "..."
            bin_labels.append(shortened_name)

    # Plot Count (%) as bars with improved styling
    bars = ax1.bar(x_indices, df['Count (%)'], color=bar_color, alpha=0.85, edgecolor='white', linewidth=0.5)

    # Add value labels on top of bars if requested
    if show_bar_values:
        for i, bar in enumerate(bars):
            height = bar.get_height()
            if height > 0:
                ax1.text(
                    bar.get_x() + bar.get_width()/2.,
                    height + 0.3,
                    f'{height:.1f}%',
                    ha='center', va='bottom',
                    fontsize=8, fontweight='bold', color=bar_color,
                    bbox=dict(boxstyle='round,pad=0.2', facecolor='white', edgecolor='none', alpha=0.7)
                )

    ax1.set_xticks(x_indices)
    ax1.set_xticklabels(bin_labels, rotation=45, ha='right', fontsize=9)
    ax1.set_xlabel('Bins', fontsize=10, fontweight='bold')
    ax1.set_ylabel('Count (%)', color=bar_color, fontsize=10, fontweight='bold')
    ax1.tick_params(axis='y', labelcolor=bar_color, labelsize=9)

    # Create second y-axis for Event Rate or WoE
    ax2 = ax1.twinx()

    # Label style for line annotations
    label_bbox = dict(boxstyle='round,pad=0.2', facecolor='white', edgecolor='none', alpha=0.8)

    # Plot the metric line for regular bins only
    if not regular_bins.empty:
        regular_mask = ~df[bin_col_name].str.contains('Special|Missing', regex=True, na=False)
        regular_indices = np.where(regular_mask)[0]

        if len(regular_indices) > 0:
            ax2.plot(regular_indices, regular_bins[metric_column],
                     marker='o', color=line_color, linewidth=2.5, markersize=6,
                     label=f'{dataset_name} {y_axis_label}')

            # Add value annotations with background
            for idx, val in zip(regular_indices, regular_bins[metric_column]):
                ax2.annotate(f'{val:.2f}',
                             xy=(idx, val),
                             xytext=(0, 8),
                             textcoords='offset points',
                             ha='center',
                             fontsize=8, fontweight='bold', color=line_color,
                             bbox=label_bbox)

    # Plot Special bin point (if exists and has data)
    if not special_bin.empty:
        special_indices = df[df[bin_col_name].str.contains('Special', regex=False, na=False)].index
        for idx in special_indices:
            special_val = df.loc[idx, metric_column]
            if special_val != 0:
                ax2.plot(idx, special_val,
                         marker='s', color=CHART_COLORS['train_special'], markersize=8, linestyle='None',
                         label='Special' if idx == special_indices[0] else "")
                ax2.annotate(f'{special_val:.2f}',
                             xy=(idx, special_val),
                             xytext=(0, 8),
                             textcoords='offset points',
                             ha='center',
                             fontsize=8, fontweight='bold', color=CHART_COLORS['train_special'],
                             bbox=label_bbox)

    # Plot Missing bin point (if exists and has data)
    if not missing_bin.empty:
        missing_indices = df[df[bin_col_name].str.contains('Missing', regex=False, na=False)].index
        for idx in missing_indices:
            missing_val = df.loc[idx, metric_column]
            if missing_val != 0:
                ax2.plot(idx, missing_val,
                         marker='D', color=CHART_COLORS['train_missing'], markersize=8, linestyle='None',
                         label='Missing' if idx == missing_indices[0] else "")
                ax2.annotate(f'{missing_val:.2f}',
                             xy=(idx, missing_val),
                             xytext=(0, 8),
                             textcoords='offset points',
                             ha='center',
                             fontsize=8, fontweight='bold', color=CHART_COLORS['train_missing'],
                             bbox=label_bbox)

    # Add horizontal line for Totals if available
    if totals_row is not None:
        total_metric_value = totals_row[metric_column]

        # Handle case where value might be a blank string
        if isinstance(total_metric_value, str) and total_metric_value.strip() == '':
            total_metric_value = 0.0
        else:
            try:
                total_metric_value = float(total_metric_value)
            except (ValueError, TypeError):
                total_metric_value = 0.0

        ax2.axhline(y=total_metric_value, color=line_color, linestyle='--',
                   alpha=0.7, linewidth=1.5, label=f'{dataset_name} Total: {total_metric_value:.2f}')

    # Set y-axis label
    ax2.set_ylabel(y_axis_label, color=line_color, fontsize=11, fontweight='bold')
    ax2.tick_params(axis='y', labelcolor=line_color)

    # Add legend with unique entries
    handles, labels = ax2.get_legend_handles_labels()
    by_label = dict(zip(labels, handles))
    ax2.legend(by_label.values(), by_label.keys(), loc='upper right', fontsize=9, framealpha=0.9)


def plot_comparison(
    ax1, train_df, oot_df, bin_col_name, metric_column,
    y_axis_label, train_totals_row, oot_totals_row, metric,
    show_bar_values=False
) -> None:
    """
    Helper function to plot a comparison between training and OOT datasets

    Parameters:
    -----------
    ax1 : matplotlib axis
        The primary axis for plotting
    train_df : pandas DataFrame
        Training data binning table
    oot_df : pandas DataFrame
        OOT data binning table
    bin_col_name : str
        Name of the bin column
    metric_column : str
        Name of the metric column ('Event rate' or 'WoE')
    y_axis_label : str
        Label for the y-axis
    train_totals_row : pandas Series
        Totals row from training data
    oot_totals_row : pandas Series
        Totals row from OOT data
    metric : str
        Metric to plot ('event_rate' or 'woe')
    show_bar_values : bool
        Whether to show Count (%) values on top of the bars
    """
    # Ensure bin columns are string type for consistency
    train_df[bin_col_name] = train_df[bin_col_name].astype(str)
    oot_df[bin_col_name] = oot_df[bin_col_name].astype(str)

    # Create a combined set of unique bin names, preserving order from train_df first
    bin_names = list(train_df[bin_col_name])
    # Add any bins from OOT that aren't in training (shouldn't happen normally)
    for bin_name in oot_df[bin_col_name]:
        if bin_name not in bin_names:
            bin_names.append(bin_name)

    # Create x indices
    x_indices = np.arange(len(bin_names))

    # Prepare data for plotting - match bin names between datasets
    train_values = []
    oot_values = []
    train_counts = []
    oot_counts = []

    for bin_name in bin_names:
        # Find metric values for each bin in each dataset
        train_row = train_df[train_df[bin_col_name] == bin_name]
        oot_row = oot_df[oot_df[bin_col_name] == bin_name]

        # If bin exists in training data, get its value; otherwise NaN
        if not train_row.empty:
            train_values.append(train_row[metric_column].values[0])
            train_counts.append(train_row['Count (%)'].values[0])
        else:
            train_values.append(np.nan)
            train_counts.append(0)

        # If bin exists in OOT data, get its value; otherwise NaN
        if not oot_row.empty:
            oot_values.append(oot_row[metric_column].values[0])
            oot_counts.append(oot_row['Count (%)'].values[0])
        else:
            oot_values.append(np.nan)
            oot_counts.append(0)

    # Format bin labels
    bin_labels = []
    for bin_name in bin_names:
        if 'Special' in bin_name or 'Missing' in bin_name:
            bin_labels.append(bin_name)
        else:
            shortened_name = str(bin_name)
            if len(shortened_name) > 15:
                shortened_name = shortened_name[:12] + "..."
            bin_labels.append(shortened_name)

    # Plot Count (%) as grouped bars with improved colors
    bar_width = 0.35
    bar1 = ax1.bar(x_indices - bar_width/2, train_counts, bar_width,
                   color=CHART_COLORS['train_bar'], alpha=0.85,
                   edgecolor='white', linewidth=0.5, label='Train Count (%)')
    bar2 = ax1.bar(x_indices + bar_width/2, oot_counts, bar_width,
                   color=CHART_COLORS['oot_bar'], alpha=0.85,
                   edgecolor='white', linewidth=0.5, label='OOT Count (%)')

    # Label style with background for readability
    bar_label_bbox = dict(boxstyle='round,pad=0.15', facecolor='white', edgecolor='none', alpha=0.7)
    line_label_bbox = dict(boxstyle='round,pad=0.2', facecolor='white', edgecolor='none', alpha=0.85)

    # Add value labels on top of bars if requested
    if show_bar_values:
        for i, (b1, b2) in enumerate(zip(bar1, bar2)):
            # Training bar values
            h1 = b1.get_height()
            if h1 > 0:
                ax1.text(
                    b1.get_x() + b1.get_width()/2.,
                    h1 + 0.3,
                    f'{h1:.1f}%',
                    ha='center', va='bottom',
                    fontsize=7, fontweight='bold', color=CHART_COLORS['train_bar'],
                    bbox=bar_label_bbox
                )

            # OOT bar values
            h2 = b2.get_height()
            if h2 > 0:
                ax1.text(
                    b2.get_x() + b2.get_width()/2.,
                    h2 + 0.3,
                    f'{h2:.1f}%',
                    ha='center', va='bottom',
                    fontsize=7, fontweight='bold', color=CHART_COLORS['oot_bar'],
                    bbox=bar_label_bbox
                )

    ax1.set_xticks(x_indices)
    ax1.set_xticklabels(bin_labels, rotation=45, ha='right', fontsize=9)
    ax1.set_xlabel('Bins', fontsize=10, fontweight='bold')
    ax1.set_ylabel('Count (%)', color=CHART_COLORS['train_bar'], fontsize=10, fontweight='bold')
    ax1.tick_params(axis='y', labelcolor=CHART_COLORS['train_bar'], labelsize=9)
    ax1.legend(loc='upper left', fontsize=8, framealpha=0.9)

    # Create second y-axis for Event Rate or WoE
    ax2 = ax1.twinx()

    # Identify regular, special, and missing bins
    regular_mask = ~np.array([('Special' in b or 'Missing' in b) for b in bin_names])
    regular_indices = np.where(regular_mask)[0]
    special_indices = [i for i, b in enumerate(bin_names) if 'Special' in b]
    missing_indices = [i for i, b in enumerate(bin_names) if 'Missing' in b]

    # Plot lines for regular bins
    if len(regular_indices) > 0:
        # Filter out NaN values for line plots
        valid_train_idx = [i for i in regular_indices if not np.isnan(train_values[i])]
        valid_oot_idx = [i for i in regular_indices if not np.isnan(oot_values[i])]

        if valid_train_idx:
            # Plot training data line
            ax2.plot([i for i in valid_train_idx],
                     [train_values[i] for i in valid_train_idx],
                     marker='o', color=CHART_COLORS['train_line'], linewidth=2.5, markersize=6,
                     label=f'Train {y_axis_label}')

            # Train line labels - positioned ABOVE
            for idx in valid_train_idx:
                ax2.annotate(f'{train_values[idx]:.2f}',
                             xy=(idx, train_values[idx]),
                             xytext=(0, 10),
                             textcoords='offset points',
                             ha='center',
                             fontsize=8, fontweight='bold',
                             color=CHART_COLORS['train_label'],
                             bbox=line_label_bbox)

        if valid_oot_idx:
            ax2.plot([i for i in valid_oot_idx],
                     [oot_values[i] for i in valid_oot_idx],
                     marker='s', color=CHART_COLORS['oot_line'], linewidth=2.5, markersize=6,
                     label=f'OOT {y_axis_label}')

            # OOT line labels - positioned BELOW
            for idx in valid_oot_idx:
                ax2.annotate(f'{oot_values[idx]:.2f}',
                             xy=(idx, oot_values[idx]),
                             xytext=(0, -14),
                             textcoords='offset points',
                             ha='center',
                             fontsize=8, fontweight='bold',
                             color=CHART_COLORS['oot_label'],
                             bbox=line_label_bbox)

    # Plot Special bin points (only if non-zero)
    for idx in special_indices:
        train_val = train_values[idx]
        oot_val = oot_values[idx]

        if not np.isnan(train_val) and train_val != 0:
            ax2.plot(idx, train_val,
                     marker='s', color=CHART_COLORS['train_special'], markersize=8, linestyle='None',
                     label='Train Special' if idx == special_indices[0] else "")
            ax2.annotate(f'{train_val:.2f}',
                         xy=(idx, train_val),
                         xytext=(-12, 6),
                         textcoords='offset points',
                         ha='center',
                         fontsize=8, fontweight='bold',
                         color=CHART_COLORS['train_special'],
                         bbox=line_label_bbox)

        if not np.isnan(oot_val) and oot_val != 0:
            ax2.plot(idx, oot_val,
                     marker='s', color=CHART_COLORS['oot_special'], markersize=8, linestyle='None',
                     label='OOT Special' if idx == special_indices[0] else "")
            ax2.annotate(f'{oot_val:.2f}',
                         xy=(idx, oot_val),
                         xytext=(12, -6),
                         textcoords='offset points',
                         ha='center',
                         fontsize=8, fontweight='bold',
                         color=CHART_COLORS['oot_special'],
                         bbox=line_label_bbox)

    # Plot Missing bin points (only if non-zero)
    for idx in missing_indices:
        train_val = train_values[idx]
        oot_val = oot_values[idx]

        if not np.isnan(train_val) and train_val != 0:
            ax2.plot(idx, train_val,
                     marker='D', color=CHART_COLORS['train_missing'], markersize=8, linestyle='None',
                     label='Train Missing' if idx == missing_indices[0] else "")
            ax2.annotate(f'{train_val:.2f}',
                         xy=(idx, train_val),
                         xytext=(-12, 6),
                         textcoords='offset points',
                         ha='center',
                         fontsize=8, fontweight='bold',
                         color=CHART_COLORS['train_missing'],
                         bbox=line_label_bbox)

        if not np.isnan(oot_val) and oot_val != 0:
            ax2.plot(idx, oot_val,
                     marker='D', color=CHART_COLORS['oot_missing'], markersize=8, linestyle='None',
                     label='OOT Missing' if idx == missing_indices[0] else "")
            ax2.annotate(f'{oot_val:.2f}',
                         xy=(idx, oot_val),
                         xytext=(12, -6),
                         textcoords='offset points',
                         ha='center',
                         fontsize=8, fontweight='bold',
                         color=CHART_COLORS['oot_missing'],
                         bbox=line_label_bbox)

    # Add horizontal lines for totals
    if train_totals_row is not None and metric_column in train_totals_row:
        try:
            train_total_val = train_totals_row[metric_column]
            if train_total_val != '' and not (isinstance(train_total_val, float) and np.isnan(train_total_val)):
                train_total_metric = float(train_total_val)
                ax2.axhline(y=train_total_metric, color=CHART_COLORS['train_line'], linestyle='--',
                           alpha=0.7, linewidth=1.5, label=f'Train Total: {train_total_metric:.2f}')
        except (ValueError, TypeError):
            pass

    if oot_totals_row is not None and metric_column in oot_totals_row:
        try:
            oot_total_val = oot_totals_row[metric_column]
            if oot_total_val != '' and not (isinstance(oot_total_val, float) and np.isnan(oot_total_val)):
                oot_total_metric = float(oot_total_val)
                ax2.axhline(y=oot_total_metric, color=CHART_COLORS['oot_line'], linestyle='--',
                           alpha=0.7, linewidth=1.5, label=f'OOT Total: {oot_total_metric:.2f}')
        except (ValueError, TypeError):
            pass

    # Set y-axis label and legend
    ax2.set_ylabel(y_axis_label, color='#2C3E50', fontsize=11, fontweight='bold')
    ax2.tick_params(axis='y', labelcolor='#2C3E50')

    # Add legend with unique entries
    handles, labels = ax2.get_legend_handles_labels()
    by_label = dict(zip(labels, handles))
    ax2.legend(by_label.values(), by_label.keys(), loc='upper right', fontsize=8, framealpha=0.9)